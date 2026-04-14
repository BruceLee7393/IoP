"""
Excel 处理工具模块
提供生成 Excel 文件的通用功能
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from io import BytesIO
from datetime import datetime


def create_excel_file(data, headers, sheet_name="Data Export"):
    """
    创建 Excel 文件
    :param data: 要导出的数据列表，每个元素是一个字典
    :param headers: 表头配置，格式为 [{'key': '字段名', 'title': '显示标题', 'width': 宽度}, ...]
    :param sheet_name: 工作表名称
    :return: Excel 文件的字节流
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")  # 白色粗体字
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 蓝色背景
    header_alignment = Alignment(horizontal="center", vertical="center")  # 居中对齐
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header['title'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = header.get('width', 15)
    
    # 写入数据
    for row_idx, item in enumerate(data, 2):  # 从第2行开始（第1行是表头）
        for col_idx, header in enumerate(headers, 1):
            # 获取字段值，支持嵌套字段（如 role.role_name）
            value = get_nested_value(item, header['key'])
            # 清理 Excel 不允许的控制字符
            if isinstance(value, str):
                value = ILLEGAL_CHARACTERS_RE.sub('', value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # 文本对齐
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 将工作簿保存到字节流中
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def create_order_excel_file(orders, sheet_name="order_export"):
    """
    创建订单专用的 Excel 文件，支持复杂的层次结构展示
    :param orders: 订单数据列表
    :param sheet_name: 工作表名称
    :return: Excel 文件的字节流
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 定义表头（添加序号列）
    headers = [
        "序号", "订单号", "机型", "料号", "序列号", "组件名称", "组件料号", 
        "名称", "料号", "规格型号", "软件名称", "软件版本号"
    ]
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据单元格样式
    data_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # 设置列宽
        column_letter = get_column_letter(col_idx)
        if header == "序号":
            ws.column_dimensions[column_letter].width = 8
        elif header in ["订单号", "机型", "料号", "序列号", "组件名称", "组件料号"]:
            ws.column_dimensions[column_letter].width = 25
        elif header in ["名称", "规格型号"]:
            ws.column_dimensions[column_letter].width = 25
        elif header in ["软件名称", "软件版本号"]:
            ws.column_dimensions[column_letter].width = 35
        else:
            ws.column_dimensions[column_letter].width = 20
    
    # 写入数据
    current_row = 2
    
    for order_index, order in enumerate(orders, 1):
        # 构建序列号范围字符串
        serial_range = f"{order.serial_number_start}--{order.serial_number_end}"
        
        # 获取订单的组件
        components = getattr(order, 'components', [])
        
        # 计算订单总行数
        order_total_rows = _calculate_order_rows(order)
        order_start_row = current_row
        
        if order_total_rows == 0:
            # 如果没有组件，只写订单基本信息
            row_data = [
                order_index, order.order_number, order.model, order.part_number, serial_range,
                "", "", "", "", "", "", ""
            ]
            _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
            # 序号列居中
            ws.cell(row=current_row, column=1).alignment = center_alignment
            current_row += 1
        else:
            # 遍历组件
            for component in components:
                sub_components = getattr(component, 'sub_components', [])
                
                # 计算组件总行数
                component_total_rows = _calculate_component_rows(component)
                component_start_row = current_row
                
                if not sub_components:
                    # 如果组件没有子组件，写组件信息
                    row_data = [
                        "", "", "", "", "",  # 序号和订单信息稍后合并
                        component.component_name, component.component_part_number,
                        "", "", "", "", ""
                    ]
                    _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                    current_row += 1
                else:
                    # 遍历子组件
                    for sub_component in sub_components:
                        softwares = getattr(sub_component, 'softwares', [])
                        
                        # 计算子组件总行数
                        sub_component_total_rows = _calculate_sub_component_rows(sub_component)
                        sub_component_start_row = current_row
                        
                        if not softwares:
                            # 如果子组件没有软件，写子组件信息
                            row_data = [
                                "", "", "", "", "",  # 序号和订单信息稍后合并
                                "", "",  # 组件信息稍后合并
                                sub_component.sub_component_name,
                                sub_component.sub_component_part_number,
                                getattr(sub_component, 'specification', ''),
                                "", ""
                            ]
                            _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                            current_row += 1
                        else:
                            # 遍历软件
                            for software in softwares:
                                row_data = [
                                    "", "", "", "", "",  # 序号和订单信息稍后合并
                                    "", "",  # 组件信息稍后合并
                                    "", "", "",  # 子组件信息稍后合并
                                    software.software_name,
                                    getattr(software, 'software_version', '')
                                ]
                                _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                                current_row += 1
                        
                        # 合并子组件单元格
                        if sub_component_total_rows > 1:
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 8, 
                                               sub_component.sub_component_name, thin_border, data_alignment)
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 9, 
                                               sub_component.sub_component_part_number, thin_border, data_alignment)
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 10, 
                                               getattr(sub_component, 'specification', ''), thin_border, data_alignment)
                        elif sub_component_total_rows == 1:
                            # 只有一行时，直接设置值
                            ws.cell(row=sub_component_start_row, column=8, value=sub_component.sub_component_name)
                            ws.cell(row=sub_component_start_row, column=9, value=sub_component.sub_component_part_number)
                            ws.cell(row=sub_component_start_row, column=10, value=getattr(sub_component, 'specification', ''))
                
                # 合并组件单元格
                if component_total_rows > 1:
                    _merge_and_set_value(ws, component_start_row, current_row - 1, 6, 
                                       component.component_name, thin_border, data_alignment)
                    _merge_and_set_value(ws, component_start_row, current_row - 1, 7, 
                                       component.component_part_number, thin_border, data_alignment)
                elif component_total_rows == 1:
                    # 只有一行时，直接设置值
                    ws.cell(row=component_start_row, column=6, value=component.component_name)
                    ws.cell(row=component_start_row, column=7, value=component.component_part_number)
            
            # 合并订单单元格（包括序号）
            if order_total_rows > 1:
                _merge_and_set_value(ws, order_start_row, current_row - 1, 1, 
                                   order_index, thin_border, center_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 2, 
                                   order.order_number, thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 3, 
                                   order.model, thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 4, 
                                   order.part_number, thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 5, 
                                   serial_range, thin_border, data_alignment)
            elif order_total_rows == 1:
                # 只有一行时，直接设置值
                cell = ws.cell(row=order_start_row, column=1, value=order_index)
                cell.alignment = center_alignment
                ws.cell(row=order_start_row, column=2, value=order.order_number)
                ws.cell(row=order_start_row, column=3, value=order.model)
                ws.cell(row=order_start_row, column=4, value=order.part_number)
                ws.cell(row=order_start_row, column=5, value=serial_range)
        
        # 在每个订单后添加空行分隔
        #current_row += 2
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 将工作簿保存到字节流中
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def _write_excel_row(ws, row_num, data, border_style, alignment):
    """
    写入Excel行数据的辅助函数
    """
    for col_idx, value in enumerate(data, 1):
        # 清理 Excel 不允许的控制字符
        if isinstance(value, str):
            value = ILLEGAL_CHARACTERS_RE.sub('', value)
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = border_style
        cell.alignment = alignment


def _calculate_order_rows(order):
    """计算订单需要的总行数"""
    total_rows = 0
    components = getattr(order, 'components', [])
    
    if not components:
        return 1
    
    for component in components:
        total_rows += _calculate_component_rows(component)
    
    return total_rows


def _calculate_component_rows(component):
    """计算组件需要的总行数"""
    sub_components = getattr(component, 'sub_components', [])
    
    if not sub_components:
        return 1
    
    total_rows = 0
    for sub_component in sub_components:
        total_rows += _calculate_sub_component_rows(sub_component)
    
    return total_rows


def _calculate_sub_component_rows(sub_component):
    """计算子组件需要的总行数"""
    softwares = getattr(sub_component, 'softwares', [])
    
    if not softwares:
        return 1
    
    return len(softwares)


def _merge_and_set_value(ws, start_row, end_row, col, value, border, alignment):
    """合并单元格并设置值、样式"""
    if start_row == end_row:
        # 只有一行，不需要合并，直接设置值
        cell = ws.cell(row=start_row, column=col, value=value)
        cell.border = border
        cell.alignment = alignment
    else:
        # 多行，需要合并
        start_cell = f"{get_column_letter(col)}{start_row}"
        end_cell = f"{get_column_letter(col)}{end_row}"
        ws.merge_cells(f"{start_cell}:{end_cell}")
        
        # 设置合并后单元格的值和样式
        merged_cell = ws[start_cell]
        merged_cell.value = value
        merged_cell.border = border
        merged_cell.alignment = alignment
        
        # 为合并区域内的所有单元格设置边框
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border


def get_nested_value(obj, key):
    """
    获取嵌套对象的值
    :param obj: 对象（可能是字典或对象实例）
    :param key: 键名，支持点号分隔的嵌套键（如 'role.role_name'）
    :return: 对应的值，如果不存在返回空字符串
    """
    try:
        # 分割键名
        keys = key.split('.')
        value = obj
        
        for k in keys:
            if hasattr(value, k):
                # 如果是对象属性
                value = getattr(value, k)
            elif isinstance(value, dict) and k in value:
                # 如果是字典键
                value = value[k]
            else:
                return ""
        
        # 处理 None 值
        if value is None:
            return ""
        
        # 处理日期时间格式
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
            
        return str(value)
    except:
        return ""


def generate_filename(prefix="导出数据"):
    """
    生成带时间戳的文件名
    :param prefix: 文件名前缀
    :return: 文件名字符串
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"


# ==================== 订单批量导入功能 ====================

def parse_order_excel(file_stream):
    """
    解析订单 Excel 文件
    :param file_stream: 文件流
    :return: (success_orders, error_orders)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(BytesIO(file_stream.read()), data_only=True)
    ws = wb.active
    
    # 1. 查找表头行
    header_row_idx, column_indices = _find_header_row(ws)
    if not header_row_idx:
        raise ValueError("未找到有效的表头行，请确保 Excel 包含：序号、订单号、机型等列")
    
    # 2. 解析数据行，构建订单树
    orders = _parse_data_rows(ws, header_row_idx + 1, column_indices)
    
    # 3. 验证订单数据
    success_orders = []
    error_orders = []
    
    for order in orders:
        error = _validate_order(order)
        if error:
            error_orders.append({
                'order': order,
                'error': error
            })
        else:
            success_orders.append(order)
    
    return success_orders, error_orders


def _find_header_row(ws):
    """查找表头行"""
    key_columns = {
        'order_index': ['序号', '编号'],
        'order_number': ['订单号'],
        'model': ['机型', '型号'],
        'part_number': ['料号'],
        'serial_number': ['序列号'],
        'component_name': ['组件名称'],
        'component_part_number': ['组件料号'],
        'sub_component_name': ['子件名称', '子组件名称', '名称'],
        'sub_component_part_number': ['子件料号', '子组件料号'],
        'specification': ['规格型号', '规格'],
        'software_name': ['软件名称'],
        'software_version': ['软件版本号', '版本号']
    }
    
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        cleaned_row = [_clean_cell_value(cell) for cell in row]
        
        if '订单号' in cleaned_row and '机型' in cleaned_row:
            column_indices = {}
            for key, variants in key_columns.items():
                for variant in variants:
                    if variant in cleaned_row:
                        column_indices[key] = cleaned_row.index(variant)
                        break
            
            required = ['order_number', 'model', 'part_number', 'serial_number']
            if all(key in column_indices for key in required):
                return row_idx, column_indices
    
    return None, None


def _parse_data_rows(ws, start_row, col_idx):
    """解析数据行，构建订单树结构"""
    orders = []
    current_order = None
    current_component = None
    current_sub_component = None
    
    for row_idx in range(start_row, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        if not any(row):
            continue
        
        order_index = _get_cell_value(row, col_idx, 'order_index')
        order_number = _get_cell_value(row, col_idx, 'order_number')
        model = _get_cell_value(row, col_idx, 'model')
        part_number = _get_cell_value(row, col_idx, 'part_number')
        serial_number = _get_cell_value(row, col_idx, 'serial_number')
        component_name = _get_cell_value(row, col_idx, 'component_name')
        component_part_number = _get_cell_value(row, col_idx, 'component_part_number')
        sub_component_name = _get_cell_value(row, col_idx, 'sub_component_name')
        sub_component_part_number = _get_cell_value(row, col_idx, 'sub_component_part_number')
        specification = _get_cell_value(row, col_idx, 'specification')
        software_name = _get_cell_value(row, col_idx, 'software_name')
        software_version = _get_cell_value(row, col_idx, 'software_version')
        
        # 新订单
        if order_number:
            serial_start, serial_end = _parse_serial_number(serial_number)
            current_order = {
                'order_index': order_index,
                'order_number': order_number,
                'model': model,
                'part_number': part_number,
                'serial_number_start': serial_start,
                'serial_number_end': serial_end,
                'components': []
            }
            orders.append(current_order)
            current_component = None
            current_sub_component = None
        
        # 新组件
        if component_name and current_order:
            current_component = {
                'component_name': component_name,
                'component_part_number': component_part_number,
                'sub_components': []
            }
            current_order['components'].append(current_component)
            current_sub_component = None
        
        # 新子组件
        if sub_component_name and current_component:
            current_sub_component = {
                'sub_component_name': sub_component_name,
                'sub_component_part_number': sub_component_part_number,
                'specification': specification,
                'softwares': []
            }
            current_component['sub_components'].append(current_sub_component)
        
        # 添加软件
        if software_name and current_sub_component:
            current_sub_component['softwares'].append({
                'software_name': software_name,
                'software_version': software_version
            })
    
    return orders


def _validate_order(order):
    """验证订单数据完整性,返回所有错误信息列表"""
    errors = []
    
    if not order.get('order_number'):
        errors.append('订单号不能为空')
    if not order.get('model'):
        errors.append('机型不能为空')
    if not order.get('part_number'):
        errors.append('料号不能为空')
    if not order.get('serial_number_start'):
        errors.append('序列号起始不能为空')
    if not order.get('serial_number_end'):
        errors.append('序列号结束不能为空')
    
    components = order.get('components', [])
    if not components:
        errors.append('订单必须包含至少一个组件')
    
    for comp in components:
        comp_name = comp.get('component_name', '未知')
        if not comp.get('component_name'):
            errors.append("组件名称不能为空")
        if not comp.get('component_part_number'):
            errors.append(f"组件料号不能为空（组件：{comp_name}）")
        
        sub_components = comp.get('sub_components', [])
        if not sub_components:
            errors.append(f"组件必须包含至少一个子组件（组件：{comp_name}）")
        
        for sub_comp in sub_components:
            sub_comp_name = sub_comp.get('sub_component_name', '未知')
            if not sub_comp.get('sub_component_name'):
                errors.append(f"子组件名称不能为空（组件：{comp_name}）")
            if not sub_comp.get('sub_component_part_number'):
                errors.append(f"子组件料号不能为空（子组件：{sub_comp_name}）")
            
            softwares = sub_comp.get('softwares', [])
            if not softwares:
                errors.append(f"子组件必须包含至少一个软件（子组件：{sub_comp_name}）")
            
            for sw in softwares:
                if not sw.get('software_name'):
                    errors.append(f"软件名称不能为空（子组件：{sub_comp_name}）")
    
    return '; '.join(errors) if errors else None


def _parse_serial_number(serial_str):
    """解析序列号范围"""
    if not serial_str:
        return '', ''
    
    serial_str = serial_str.replace('\n', '--').replace('\r', '').strip()
    
    if '--' in serial_str:
        parts = serial_str.split('--', 1)
    elif '-' in serial_str and serial_str.count('-') >= 2:
        parts = serial_str.rsplit('-', 1)
    else:
        parts = [serial_str]
    
    if len(parts) >= 2:
        return parts[0].strip(), parts[1].strip()
    return parts[0].strip(), parts[0].strip()


def _get_cell_value(row, col_idx, key):
    """安全获取单元格值"""
    if key not in col_idx or col_idx[key] >= len(row):
        return ''
    return _clean_cell_value(row[col_idx[key]])


def _clean_cell_value(value):
    """清理单元格值"""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.replace('\ufeff', '').replace('\u200b', '').strip()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(int(value)) if abs(value) >= 1e10 else str(value)
    return str(value).strip()


def create_error_excel(error_orders):
    """创建错误订单 Excel 文件（保持与导出相同的格式，带合并单元格）"""
    wb = Workbook()
    ws = wb.active
    ws.title = '导入错误'
    
    headers = ['序号', '订单号', '机型', '料号', '序列号', '组件名称', '组件料号', 
               '名称', '料号', '规格型号', '软件名称', '软件版本号', '错误原因']
    
    # 样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    error_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        column_letter = get_column_letter(col)
        if col == 1:  # 序号
            ws.column_dimensions[column_letter].width = 8
        elif col == 13:  # 错误原因
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = 25
    
    current_row = 2
    
    for error_item in error_orders:
        order = error_item['order']
        error_msg = error_item['error']
        order_index = order.get('order_index', '')
        
        # 构建序列号
        serial = f"{order.get('serial_number_start', '')}--{order.get('serial_number_end', '')}"
        
        # 获取组件
        components = order.get('components', [])
        order_start_row = current_row
        
        if not components:
            # 没有组件，只写订单信息
            row_data = [order_index, order.get('order_number', ''), order.get('model', ''),
                       order.get('part_number', ''), serial, '', '', '', '', '', '', '', error_msg]
            _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
            ws.cell(current_row, 1).alignment = center_alignment
            ws.cell(current_row, 13).fill = error_fill
            current_row += 1
        else:
            # 遍历组件
            for component in components:
                sub_components = component.get('sub_components', [])
                component_start_row = current_row
                
                if not sub_components:
                    # 组件没有子组件
                    row_data = ['', '', '', '', '', component.get('component_name', ''),
                               component.get('component_part_number', ''), '', '', '', '', '', '']
                    _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                    current_row += 1
                else:
                    # 遍历子组件 - 每个子组件必须写入
                    for sub_component in sub_components:
                        softwares = sub_component.get('softwares', [])
                        sub_component_start_row = current_row
                        
                        # 每个子组件至少写一行（即使没有软件）
                        if not softwares:
                            # 子组件没有软件，写一行空软件
                            row_data = ['', '', '', '', '', '', '',
                                       sub_component.get('sub_component_name', ''),
                                       sub_component.get('sub_component_part_number', ''),
                                       sub_component.get('specification', ''), '', '', '']
                            _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                            current_row += 1
                        else:
                            # 有软件，每个软件一行
                            for software in softwares:
                                row_data = ['', '', '', '', '', '', '', '', '', '',
                                           software.get('software_name', ''),
                                           software.get('software_version', ''), '']
                                _write_excel_row(ws, current_row, row_data, thin_border, data_alignment)
                                current_row += 1
                        
                        # 合并子组件单元格
                        sub_rows = current_row - sub_component_start_row
                        if sub_rows > 1:
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 8,
                                               sub_component.get('sub_component_name', ''), thin_border, data_alignment)
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 9,
                                               sub_component.get('sub_component_part_number', ''), thin_border, data_alignment)
                            _merge_and_set_value(ws, sub_component_start_row, current_row - 1, 10,
                                               sub_component.get('specification', ''), thin_border, data_alignment)
                        else:
                            # 只有一行时直接设置
                            ws.cell(sub_component_start_row, 8).value = sub_component.get('sub_component_name', '')
                            ws.cell(sub_component_start_row, 9).value = sub_component.get('sub_component_part_number', '')
                            ws.cell(sub_component_start_row, 10).value = sub_component.get('specification', '')
                
                # 合并组件单元格
                comp_rows = current_row - component_start_row
                if comp_rows > 1:
                    _merge_and_set_value(ws, component_start_row, current_row - 1, 6,
                                       component.get('component_name', ''), thin_border, data_alignment)
                    _merge_and_set_value(ws, component_start_row, current_row - 1, 7,
                                       component.get('component_part_number', ''), thin_border, data_alignment)
                else:
                    # 只有一行时直接设置
                    ws.cell(component_start_row, 6).value = component.get('component_name', '')
                    ws.cell(component_start_row, 7).value = component.get('component_part_number', '')
            
            # 合并订单单元格（包括序号和错误原因）
            order_rows = current_row - order_start_row
            if order_rows > 1:
                _merge_and_set_value(ws, order_start_row, current_row - 1, 1,
                                   order_index, thin_border, center_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 2,
                                   order.get('order_number', ''), thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 3,
                                   order.get('model', ''), thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 4,
                                   order.get('part_number', ''), thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 5,
                                   serial, thin_border, data_alignment)
                _merge_and_set_value(ws, order_start_row, current_row - 1, 13,
                                   error_msg, thin_border, data_alignment)
            else:
                ws.cell(order_start_row, 1, order_index).alignment = center_alignment
                ws.cell(order_start_row, 2, order.get('order_number', ''))
                ws.cell(order_start_row, 3, order.get('model', ''))
                ws.cell(order_start_row, 4, order.get('part_number', ''))
                ws.cell(order_start_row, 5, serial)
                ws.cell(order_start_row, 13, error_msg)
            
            # 给错误原因列所有行添加红色背景
            for row in range(order_start_row, current_row):
                ws.cell(row, 13).fill = error_fill
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_template_excel():
    """
    创建导入模板 Excel（使用与导出相同的格式）
    :return: BytesIO 对象
    """
    # 使用示例数据调用导出函数生成模板
    from collections import namedtuple
    
    # 创建示例数据结构
    Software = namedtuple('Software', ['software_name', 'software_version'])
    SubComponent = namedtuple('SubComponent', ['sub_component_name', 'sub_component_part_number', 'specification', 'softwares'])
    Component = namedtuple('Component', ['component_name', 'component_part_number', 'sub_components'])
    Order = namedtuple('Order', ['order_number', 'model', 'part_number', 'serial_number_start', 'serial_number_end', 'components'])
    
    # 示例订单
    sample_orders = [
        Order(
            order_number='示例订单号',
            model='示例机型',
            part_number='示例料号',
            serial_number_start='起始序列号',
            serial_number_end='结束序列号',
            components=[
                Component(
                    component_name='示例组件',
                    component_part_number='组件料号',
                    sub_components=[
                        SubComponent(
                            sub_component_name='示例子组件',
                            sub_component_part_number='子组件料号',
                            specification='规格型号',
                            softwares=[
                                Software(software_name='软件名称1', software_version='版本号1'),
                                Software(software_name='软件名称2', software_version='版本号2')
                            ]
                        )
                    ]
                )
            ]
        )
    ]
    
    # 使用现有的导出函数生成模板
    return create_order_excel_file(sample_orders, sheet_name="订单导入模板")
